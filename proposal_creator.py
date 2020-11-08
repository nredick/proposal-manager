import datetime
import math
import os
import time
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font


#global vars
result = []
sub_D = []
sub_I = []
sub_Y = []
sub_option = []

count = 2
total = 0
#subtotal_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'), left=Side(style='thin'))
total_border = Border(top=Side(style='thin'), bottom=Side(style='double'), right=Side(style='thin'),
                      left=Side(style='thin'))
side_border = Border(right=Side(style='thin'), left=Side(style='thin'))
full_border = Border(right=Side(style='thin'), left=Side(style='thin'), top=Side(style='thin'),
                     bottom=Side(style='thin'))

specialConditions = [
    '1.  Please note that any components that are NOT considered to be Capital Improvements are subject',
    'to the prevailing sales tax rate. Appropriate taxes will be applied at the time of invoice.',
    '2. Warranty information is listed on the construction agreement.',
    '3.  Installation of irrigation systems, electrical components, plumbing, barbeques, annuals, '
    'etc… that may be shown',
    'on the accompanying plan are not included in this proposal unless expressly mentioned.',
    '4.  Plant materials are subject to change due to availability changes from our suppliers.',
    '5.  Redbud Development and its subcontractors take every effort to prevent the growth of weeds in '
    'newly seeded',
    'lawn areas or plant beds. There is no guarantee against the presence of weeds in newly seeded or '
    'hydroseeded areas.',
    'In addition, Redbud cannot guarantee a "sod-like" final result on seeded areas; the owner takes '
    'responsibility for',
    'cultivating the new lawn to its full potential as this may take several applications of fertilizer and '
    'several growing',
    'seasons to achieve the thickness of sod.',
    '6. Budget numbers listed above are to be considered rough estimates by phase.  All phases are '
    'co-dependant and',
    'not to be considered as individual “choices.”',
    '7. Permit fees from the local municipality or governing authority are the responsibility of the owner '
    'and will be',
    'billed directly to the owner.',
    '8.  Power and/ or water necessary to implement any work to be provided by the owner.',
    '9. Client shall provide free access to work areas for employees and vehicles and agrees to keep '
    'driveways',
    'clear and available for movement and parking of trucks and equipment during work hours.',
    '10. Unless otherwise specified in the agreement, we shall not be expected to keep gates closed for '
    'animals or',
    'children.',
    '11. Spot elevations shown on the accompanying plans are approximate only. Redbud will field verify all '
    'final',
    'proposed elevations prior to construction and reserves the right to modify grades in the field as '
    'necessary to',
    'accommodate the project needs.',
    '12. SketchUp drawings are not true or accurate representations of the project and should not be '
    'considered',
    'as such. Elements shown on these renderings are a suggestion of the project only.', '', '',
    'I have read the above conditions: ____________________________________________',
    '                                                       (Client Signature)']
phase_names = ['Project Prep & Demolition Phase', 'Sub-Surface Construction', 'Hardscape Construction',
               'Softscape Construction', 'Completion']


# helper functions
def print_conditions(proposal):
    global specialConditions, count
    count += 3
    proposal.cell(row=count, column=1).value = 'Special Conditions'
    proposal[f'A{count}'].font = Font(bold=True, underline='single')
    count += 1
    for item in specialConditions:
        proposal.cell(row=count, column=1).value = item
        count += 1


def int_to_roman(num):  # function to convert phase numbers to roman numerals
    if not isinstance(num, type(1)):
        raise TypeError("expected integer, got %s" % type(num))
    if not 0 < num < 4000:
        raise ValueError("Argument must be between 1 and 3999")
    ints = (1000, 900, 500, 400, 100, 90, 50, 40, 10, 9, 5, 4, 1)
    nums = ('M', 'CM', 'D', 'CD', 'C', 'XC', 'L', 'XL', 'X', 'IX', 'V', 'IV', 'I')
    result = []
    for n in range(len(ints)):
        c = int(num / ints[n])
        result.append(nums[n] * c)
        num -= ints[n] * c
    return ''.join(result)


def set_indices(start_index, end_index, items, item_descriptions, item_prices,
                is_option_ls):  # function to set global sub interval indices for each phase
    global sub_D, sub_I, sub_Y, sub_option, specialConditions, count
    sub_D = items[start_index:end_index]
    sub_I = item_descriptions[start_index:end_index]
    sub_Y = item_prices[start_index:end_index]
    sub_option = is_option_ls[start_index:end_index]


#main function


def create_proposal(filename):
    global result, sub_D, sub_I, sub_Y, sub_option, count, total, full_border, total_border, side_border, specialConditions, phase_names, is_option_ls, contractor, numbers, item_prices, item_descriptions, items, revision_dates, siding_finish_index
    #result.append(f'{time.ctime()} Starting proposal creation...')

    os.system(f'cp \"{filename}\" \"{filename}_temp.xlsx\"')  # make a temporary copy that can be opened as data only
    result.append(f'{time.ctime()} Created temporary file: \"{filename}_temp.xlsx\"')

    workbook_data = openpyxl.load_workbook(filename, data_only=False)
    workbook = openpyxl.load_workbook(f'{filename}_temp.xlsx', data_only=True)  # reopen the temp wb

    result.append(f'{time.ctime()} Loading reference sheet.')
    ref = workbook[workbook.sheetnames[0]]  # get the reference sheet

    col_names = [[cell.value, cell.column_letter] for cell in ref['1'] if
                 cell.value is not None]  # col letter from names in first row

    if len(col_names) != 7:
        result.append(f'{time.ctime()} ERROR: Missing a header on the budget sheet. Make sure you are using the correct template and that it is the first sheet in the Excel workbook.')
        return result

    result.append(f'{time.ctime()} Getting budget data.')
    # cells from columns in the rough budget sheet
    for couple in col_names:  # item names
        if couple[0] == 'ITEM':
            items = [x.value for x in ref[f'{couple[1]}']]

    for couple in col_names:  # item descriptions
        if couple[0] == 'DESCRIPTION':
            item_descriptions = [x.value for x in ref[f'{couple[1]}']]

    for couple in col_names:  # item prices
        if couple[0] == 'PRICE':
            item_prices = [x.value for x in ref[f'{couple[1]}']]

    for couple in col_names:  # item is an option
        if couple[0] == 'OPTIONS':
            is_option_ls = [x.value for x in ref[f'{couple[1]}']]

    for couple in col_names:  # contractor/other header info; WAS COL_E
        if couple[0] == 'CONTRACTOR':
            contractor = [x.value for x in ref[f'{couple[1]}']]

    for couple in col_names:  # COL_G = [x.value for x in ref['G']]
        if couple[0] == 'NUMBERS':
            numbers = [x.value for x in ref[f'{couple[1]}']]

    for couple in col_names:  # revision dates
        if couple[0] == 'REVISION DATES':
            revision_dates = [x.value for x in ref[f'{couple[1]}']]

    # convert the options list to true/false
    result.append(f'{time.ctime()} Get options.')
    is_option = []
    for opt in is_option_ls:
        try:
            o = opt.lower()
            if o != '':
                is_option.append(True)
            else:
                is_option.append(False)
        except AttributeError:
            is_option.append(False)

    # add proposal sheet next to the budget sheet in both wbs
    result.append(f'{time.ctime()} Add proposal sheet to workbook.')
    proposal = workbook_data.create_sheet(f'Budget', 1)

    # formatting col widths
    result.append(f'{time.ctime()} Formatting proposal.')
    proposal.column_dimensions['A'].width = 80.5
    proposal.column_dimensions['B'].width = 12.5
    proposal.column_dimensions['C'].width = 12.5
    proposal.column_dimensions['D'].width = 12.5

    proposal.cell(row=count, column=5).value = 'Item cost markups'

    result.append(f'{time.ctime()} Writing data to proposal.')
    # add initial issue date to the proposal
    issue_date = str(numbers[contractor.index('ISSUE DATE:')])
    proposal.cell(row=count, column=1).value = issue_date
    count += 1

    # now get the revision dates
    header_start = items.index('CLIENT NAME:')
    header_finish = items.index('COMPONENT')

    dates = [x.strip() for x in revision_dates[header_start:header_finish] if x is not None and type(x) is str]
    for date in dates:
        proposal.cell(row=count, column=1).value = f'     {date}'
        count += 1
    count += 1

    # add header values
    proposal.cell(row=count, column=1).value = str(items[header_start + 1]).title()
    proposal[f'A{count}'].font = Font(bold=True)
    count += 1
    proposal.cell(row=count, column=1).value = str(items[items.index('ADDRESS:') + 1]).title()
    count += 1
    try:
        address = str(items[items.index('CITY/ STATE:') + 1]).split(',')
        try:
            proposal.cell(row=count, column=1).value = str(address[0]).title() + ', ' + str(address[1]).upper()
        except IndexError:
            proposal.cell(row=count, column=1).value = items[items.index('CITY/ STATE:') + 1]
    except ValueError:
        proposal.cell(row=count, column=1).value = items[items.index('CITY/ STATE:') + 1]
    count += 2
    proposal.cell(row=count, column=1).value = 'Rough Landscape Budget'
    proposal[f'A{count}'].font = Font(bold=True)
    count += 1
    prop_num = str(numbers[contractor.index('PROPOSAL #:')])
    proposal.cell(row=count, column=1).value = f'Proposal # - P. {prop_num}'
    count += 1
    plan_num = numbers[contractor.index('PLAN #:')]
    proposal.cell(row=count, column=1).value = f'Plan # {plan_num}'
    count += 1
    siding_index = count
    proposal.cell(row=count, column=2).value = 'BASE PRICE'
    proposal[f'B{count}'].border = full_border
    proposal.cell(row=count, column=3).value = 'OPTIONS'
    proposal[f'C{count}'].border = full_border
    count += 1

    # get all items, descriptions, and prices
    to_write = []
    for i, phase in enumerate(phase_names):
        to_write.append(phase)
        try:
            set_indices(items.index(phase.upper()), items.index(phase_names[i + 1].upper()), items, item_descriptions,
                        item_prices, is_option_ls)
        except IndexError:
            set_indices(items.index(phase.upper()), items.index('BLANK ON PURPOSE', items.index(phase.upper())), items,
                        item_descriptions, item_prices, is_option_ls)
        for index in range(len(sub_D)):
            if sub_D[index] is not None and 'NOTES' not in sub_D[index] and 'BLANK ON PURPOSE' != sub_D[index] and \
                    sub_I[index] is not None:
                try:
                    #print(sub_Y[index])
                    cost = int(math.ceil((float(sub_Y[index]) / 5.0)) * 5)
                except ValueError:
                    cost = sub_Y[index]
                except TypeError:
                    cost = sub_Y[index]
                to_write.append(
                    [sub_D[index].strip().title(), sub_I[index].strip().capitalize(), cost, sub_option[index]])
        to_write.append('\n')

    # write items/des/prices to the proposal

    base_sub_cells = []
    opt_sub_cells = []
    sub_rows = []
    phase_num = 1
    phase_start = count
    for i, note in enumerate(to_write):
        if note in phase_names and to_write[i + 1]:
            phase_start = count
            if to_write[i + 1] != '\n':  # there are items in the section
                proposal.cell(row=count, column=1).value = f'{int_to_roman(phase_num)}. {note}'
                phase_num += 1
                proposal[f'A{count}'].font = Font(bold=True)
                count += 1
        elif note == '\n':  # new line indicator
            if i == len(to_write) - 1:  # end of to_write
                proposal.cell(row=count, column=1).value = 'Subtotal'
                sub_rows.append(count)
                proposal[f'B{count}'].value = f'=SUM(B{phase_start}:B{count - 1})'  # base price
                proposal[f'C{count}'].value = f'=SUM(C{phase_start}:C{count - 1})'  # options
                base_sub_cells.append(f'B{count}')
                opt_sub_cells.append(f'C{count}')
                proposal[f'A{count}'].alignment = Alignment(horizontal="right")
                proposal.cell(row=count, column=4).value = f'=SUM(D{siding_index}:D{count - 1})'
                count += 2

                proposal.cell(row=count, column=1).value = 'Total'
                base_subs = '+'.join(f'B{x}' for x in sub_rows)
                proposal.cell(row=count, column=2).value = f'=SUM({base_subs})'
                opt_subs = '+'.join(f'C{x}' for x in sub_rows)

                proposal.cell(row=count, column=3).value = f'=SUM({opt_subs})'
                proposal[f'A{count}'].alignment = Alignment(horizontal="right")
                proposal[f'B{count}'].border = total_border
                proposal[f'C{count}'].border = total_border
                proposal[f'A{count}'].font = Font(bold=True)
                siding_finish_index = count
                count += 2
            else:
                proposal.cell(row=count, column=1).value = 'Subtotal'
                sub_rows.append(count)
                proposal[f'A{count}'].alignment = Alignment(horizontal="right")
                proposal[f'B{count}'].value = f'=SUM(B{phase_start}:B{count - 1})'  # base price
                proposal[f'C{count}'].value = f'=SUM(C{phase_start}:C{count - 1})'  # options
                base_sub_cells.append(f'B{count}')
                opt_sub_cells.append(f'C{count}')
                count += 2
        else:  # is a proposal item
            #index = count
            proposal.cell(row=count, column=1).value = note[0]  # item name
            proposal.cell(row=count, column=5).value = 35  # cost multiplier
            #proposal.cell(row=count, column=4).value = f'=(1+E{count}/100)*{(note[2] * 100) / 135}'

            try:
                proposal.cell(row=count, column=4).value = f'=(1+E{count}/100)*{(note[2] * 100) / 135}'  # item price
            except TypeError:
                proposal.cell(row=count, column=4).value = 'Error'

            count += 1

            for string in note[1].split('///'):  # escape char '///' for new lines
                if len(string.strip()) > 0:
                    proposal.cell(row=count, column=1).value = f'     - {string.strip().capitalize()}'
                    proposal[f'A{count}'].alignment = Alignment(wrap_text=True)
                    count += 1
            count += 1

    # special conditions and signature
    print_conditions(proposal)
    result.append(f'{time.ctime()} Finished writing data to proposal.')

    # $ formatting
    for cell in proposal['D']:
        cell.number_format = '$###,##0.00'

    for i, cell in enumerate(proposal['C']):
        cell.number_format = '$###,##0.00'
        if siding_index < i + 1 < siding_finish_index:
            cell.border = side_border

    for i, cell in enumerate(proposal['B']):
        cell.number_format = '$###,##0.00'
        if siding_index < i + 1 < siding_finish_index:
            cell.border = side_border

    # border formatting
    for cell in base_sub_cells:
        proposal[cell].border = full_border

    for cell in opt_sub_cells:
        proposal[cell].border = full_border

    # set print area
    proposal.print_area = f'A1:C{count + 2}'
    proposal.sheet_properties.pageSetUpPr.fitToPage = True
    proposal.page_setup.fitToHeight = False

    # save/delete the finalized workbook/temp workbook
    result.append(f'{time.ctime()} Saving final workbook and deleting temporary copy.')
    workbook_data.save(filename)
    workbook.save(f'{filename}_temp.xlsx')
    os.system(f'rm \"{filename}_temp.xlsx\"')

    #workbook.save(filename)
    return result
